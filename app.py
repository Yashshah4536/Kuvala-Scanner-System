from flask import Flask, render_template, request, jsonify
import qrcode
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
import json

app = Flask(__name__)
QR_FOLDER = os.path.join("static", "qrcodes")
EXCEL_FILE = "data.xlsx"

# Ensure QR folder exists
os.makedirs(QR_FOLDER, exist_ok=True)

# Initialize Excel if not exists
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Timestamp", "ID", "Name", "Email", "Phone", "QR Filename", "Scanned Data"])
    wb.save(EXCEL_FILE)

# ------------------ QR Generator ------------------ #
@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        user_id = request.form["user_id"]
        name = request.form["name"]
        email = request.form["email"]
        phone = request.form["phone"]

        # QR Data as JSON
        qr_dict = {"ID": user_id, "Name": name, "Email": email, "Phone": phone}
        qr_data = json.dumps(qr_dict)

        filename = f"{user_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.png"
        filepath = os.path.join(QR_FOLDER, filename)

        # Generate QR code
        qr = qrcode.QRCode(
            version=1, error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10, border=4
        )
        qr.add_data(qr_data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        img.save(filepath)

        # Save details in Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), user_id, name, email, phone, filename, ""])
        wb.save(EXCEL_FILE)

        return render_template("index.html", qr_image=filepath, user_id=user_id, name=name)

    return render_template("index.html", qr_image=None)

# ------------------ QR Scanner Page ------------------ #
@app.route("/scan")
def scan():
    return render_template("scan.html")

# ------------------ Save Scanned Data ------------------ #
@app.route("/save_scan", methods=["POST"])
def save_scan():
    scanned_data = request.json.get("data", "")

    user_id = name = email = phone = ""

    # Try to parse JSON
    try:
        data_dict = json.loads(scanned_data)
        user_id = data_dict.get("ID", "")
        name = data_dict.get("Name", "")
        email = data_dict.get("Email", "")
        phone = data_dict.get("Phone", "")
    except:
        # Fallback: parse as plain text
        for line in scanned_data.splitlines():
            if line.startswith("ID:"):
                user_id = line.replace("ID:", "").strip()
            elif line.startswith("Name:"):
                name = line.replace("Name:", "").strip()
            elif line.startswith("Email:"):
                email = line.replace("Email:", "").strip()
            elif line.startswith("Phone:"):
                phone = line.replace("Phone:", "").strip()

    # Save in Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        user_id, name, email, phone, "", scanned_data
    ])
    wb.save(EXCEL_FILE)

    return jsonify({"status": "success", "message": "Scanned data saved."})

if __name__ == "__main__":
    app.run(debug=True)
